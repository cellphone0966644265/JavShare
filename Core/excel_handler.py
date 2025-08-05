# /core/excel_handler.py
import pandas as pd
import openpyxl
import tempfile
import shutil
import os
from threading import Lock

# Lock để tránh xung đột khi nhiều thread cùng ghi vào file Excel
lock = Lock()

def _load_workbook_and_sheet(file_path, sheet_name, table_name):
    """Hàm nội bộ để mở workbook, sheet và kiểm tra sự tồn tại của bảng."""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File không tồn tại: {file_path}")
    
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet '{sheet_name}'")
    
    ws = wb[sheet_name]
    if table_name not in ws.tables:
        raise ValueError(f"Không tìm thấy bảng '{table_name}'")
        
    return wb, ws, ws.tables[table_name]

def read_table(file_path, sheet_name, table_name):
    """Đọc toàn bộ dữ liệu từ một bảng trong Excel và trả về list của dictionary."""
    with lock:
        _, ws, table = _load_workbook_and_sheet(file_path, sheet_name, table_name)
        data_range = ws[table.ref]
        rows = list(data_range)
        header = [cell.value for cell in rows[0]]
        data = [[cell.value for cell in row] for row in rows[1:]]
        df = pd.DataFrame(data, columns=header)
        return df.to_dict(orient='records')

def update_cell(file_path, sheet_name, table_name, row_index, column_name, new_value):
    """Cập nhật một ô cụ thể trong bảng, dựa vào chỉ số dòng và tên cột."""
    with lock:
        wb, ws, table = _load_workbook_and_sheet(file_path, sheet_name, table_name)
        header_range = ws[table.ref.split(':')[0]]
        columns = [cell.value for cell in header_range[0]]
        
        if column_name not in columns:
            raise ValueError(f"Không tìm thấy cột '{column_name}' trong bảng.")

        col_index = columns.index(column_name) + table.ref_range.min_col
        # +2 vì: +1 để bỏ qua header của bảng, +1 nữa vì row_index bắt đầu từ 0
        actual_row = table.ref_range.min_row + row_index + 1
        ws.cell(row=actual_row, column=col_index, value=new_value)
        
        # Lưu an toàn vào file tạm rồi di chuyển để tránh mất dữ liệu
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', prefix='excel_') as tmp:
            wb.save(tmp.name)
        shutil.move(tmp.name, file_path)
        return True

def write_row(file_path, sheet_name, table_name, data_dict):
    """Ghi một dòng mới (dạng dictionary) vào cuối của một bảng."""
    with lock:
        wb, ws, table = _load_workbook_and_sheet(file_path, sheet_name, table_name)
        header_range = ws[table.ref.split(':')[0]]
        columns = [cell.value for cell in header_range[0]]
        
        next_row_num = table.ref_range.max_row + 1
        
        for col_name, value in data_dict.items():
            if col_name in columns:
                col_idx = columns.index(col_name) + table.ref_range.min_col
                ws.cell(row=next_row_num, column=col_idx, value=value)

        # Mở rộng vùng tham chiếu của bảng để bao gồm dòng mới
        max_col_letter = openpyxl.utils.get_column_letter(table.ref_range.max_col)
        table.ref = f"{table.ref_range.min_col_letter}{table.ref_range.min_row}:{max_col_letter}{next_row_num}"

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx', prefix='excel_') as tmp:
            wb.save(tmp.name)
        shutil.move(tmp.name, file_path)
        return True
